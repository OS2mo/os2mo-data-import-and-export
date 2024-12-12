# import tempfile
import tempfile
import unittest
from unittest.mock import patch

import pandas as pd
import xlsxwriter
from openpyxl import load_workbook

from exporters.sql_export.sql_table_defs import Base
from exporters.sql_export.sql_table_defs import WAdresse as Adresse
from exporters.sql_export.sql_table_defs import WBruger as Bruger
from exporters.sql_export.sql_table_defs import WEngagement as Engagement
from exporters.sql_export.sql_table_defs import WEnhed as Enhed
from exporters.sql_export.sql_table_defs import WTilknytning as Tilknytning
from reports.query_actualstate import (
    XLSXExporter,
    get_engine,
    list_employees,
    list_MED_members,
    map_dynamic_class,
    merge_dynamic_classes,
    rearrange,
    sessionmaker,
    set_of_org_units,
)


class Tests_xlxs(unittest.TestCase):
    def setUp(self):
        f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self.xlsfilename = f.name
        self.data = [
            ("Navn", "Email", "Tilknytningstype", "Enhed"),
            ("Fornavn Efternavn", "email@example.com", "Formand", "Testenhed"),
        ]
        self.employees = [
            ("Navn", "Email", "Afdeling", "Stilling", "Tlf", "CPR"),
            (
                "Fornavn Efternavn",
                "email@example.com",
                "Afdeling TEST",
                "Tester",
                "00000000",
                "123456-1234",
            ),
        ]

        workbook = xlsxwriter.Workbook(self.xlsfilename)
        excel = XLSXExporter(self.xlsfilename)
        excel.add_sheet(workbook, "MED", self.data)
        excel.add_sheet(workbook, "EMP", self.employees)
        workbook.close()

    def test_read_MED(self):
        wb = load_workbook(filename=self.xlsfilename)
        ws = wb["MED"]
        # ws contains data from excel wich can be accessed with ws.colums
        # first cell in each column contains a header, second contains data
        header = [i[0].value for i in ws.columns]
        content = [i[1].value for i in ws.columns]
        self.assertEqual(header, ["Navn", "Email", "Tilknytningstype", "Enhed"])
        self.assertEqual(
            content,
            ["Fornavn Efternavn", "email@example.com", "Formand", "Testenhed"],
        )

    def test_read_EMP(self):
        wb = load_workbook(filename=self.xlsfilename)
        ws = wb["EMP"]
        # ws contains data from excel wich can be accessed with ws.colums
        # first cell in each column contains a header, second contains data
        header = [i[0].value for i in ws.columns]
        content = [i[1].value for i in ws.columns]
        self.assertEqual(
            header, ["Navn", "Email", "Afdeling", "Stilling", "Tlf", "CPR"]
        )
        self.assertEqual(
            content,
            [
                "Fornavn Efternavn",
                "email@example.com",
                "Afdeling TEST",
                "Tester",
                "00000000",
                "123456-1234",
            ],
        )


def test_rearrange():
    columns_before = [
        "Tilknytningsuuid",
        "Tilknytningens startdato",
        "Tilknytningens slutdato",
        "Navn",
        "Email",
        "Telefonnummer",
        "Tilknytningstype",
        "Tilknytningsenhed",
        "Ansættelsesenhed",
        "Enhed1",
        "Enhed2",
        "Enhed3",
        "Hovedorganisation / Faglig organisation",
    ]
    columns_after = [
        "Tilknytningens startdato",
        "Tilknytningens slutdato",
        "Navn",
        "Email",
        "Hovedorganisation / Faglig organisation",
        "Telefonnummer",
        "Tilknytningstype",
        "Tilknytningsenhed",
        "Ansættelsesenhed",
        "Enhed1",
        "Enhed2",
        "Enhed3",
    ]
    data_df = pd.DataFrame(
        [
            (
                "testuuid",
                "TEST",
                "TEST",
                "test",
                "test",
                "test",
                "test",
                "test",
                "test",
                "test",
                "test",
                "test",
                "test",
            )
        ],
        columns=columns_before,
    )
    data_df = rearrange(data_df)
    assert list(data_df.columns) == columns_after


def test_map_dynamic_class():
    returned_data = [
        {
            "uuid": "03d133d8-656a-4c8e-bffe-867b30e088a2",
            "objects": [
                {
                    "dynamic_class": {
                        "name": "Testorganisation",
                        "parent": {"name": "Hovedorganisation"},
                    }
                }
            ],
        },
        {
            "uuid": "00bffb5f-9975-4b72-a6f2-afb3ff6e5295",
            "objects": [
                {"dynamic_class": {"name": "Testorganisation", "parent": None}}
            ],
        },
        {
            "uuid": "022e7717-a023-4577-b6ee-1eec5dee63c1",
            "objects": [{"dynamic_class": None}],
        },
    ]
    expected = {
        "03d133d8-656a-4c8e-bffe-867b30e088a2": "Hovedorganisation / Testorganisation",
        "00bffb5f-9975-4b72-a6f2-afb3ff6e5295": "Testorganisation",
        "022e7717-a023-4577-b6ee-1eec5dee63c1": None,
    }

    result = map_dynamic_class(returned_data)
    assert result == expected


def test_merge_dynamic_classes():
    data_df = pd.DataFrame(
        [
            (
                "testuuid",
                "test",
                "test",
                "test",
                "test",
                "test",
                "test",
                "test",
                "test",
                "test",
            )
        ],
        columns=[
            "Tilknytningsuuid",
            "Tilknytningens startdato",
            "Tilknytningens slutdato",
            "Navn",
            "Email",
            "Telefonnummer",
            "Tilknytningstype",
            "Tilknytningsenhed",
            "Ansættelsesenhed",
            "Sti",
        ],
    )
    association_map = {"testuuid": "TestTilknytning"}
    data_df = merge_dynamic_classes(
        data_df=data_df, association_dynamic_classes=association_map
    )
    assert data_df["Hovedorganisation / Faglig organisation"][0] == "TestTilknytning"


class Tests_db(unittest.TestCase):
    def setUp(self):
        self.engine = get_engine(dbpath=":memory:")
        self.session = sessionmaker(bind=self.engine, autoflush=False)()
        # Lav tables via tabledefs fra LoraCache og fyld dataen ind
        Base.metadata.create_all(self.engine)
        enhed = Enhed(
            navn="LØN-org", uuid="LE1", bvn="Løn", enhedstype_titel="org_unit_type"
        )
        self.session.add(enhed)
        enhed = Enhed(
            navn="Under-Enhed",
            uuid="LE2",
            enhedstype_titel="org_unit_type",
            forældreenhed_uuid="LE1",
            organisatorisk_sti="LØN-org\\Under-Enhed",
            bvn="UUM",
        )
        self.session.add(enhed)
        enhed = Enhed(
            navn="Hoved-MED", uuid="E1", enhedstype_titel="org_unit_type", bvn="HM"
        )
        self.session.add(enhed)
        enhed = Enhed(
            navn="Under-MED",
            uuid="E2",
            bvn="UM",
            enhedstype_titel="org_unit_type",
            forældreenhed_uuid="E1",
        )
        self.session.add(enhed)
        enhed = Enhed(
            navn="Under-under-MED",
            uuid="E3",
            bvn="UUM",
            enhedstype_titel="org_unit_type",
            forældreenhed_uuid="E2",
        )
        self.session.add(enhed)
        bruger = Bruger(
            fornavn="fornavn",
            efternavn="efternavn",
            uuid="b1",
            bvn="b1bvn",
            cpr="cpr1",
        )
        self.session.add(bruger)
        tilknytning = Tilknytning(
            uuid="t1",
            bvn="t1bvn",
            bruger_uuid="b1",
            enhed_uuid="E2",
            tilknytningstype_titel="titel",
            startdato="2023-01-01",
            slutdato="2023-01-02",
        )
        self.session.add(tilknytning)
        engagement = Engagement(
            uuid="Eng1",
            bvn="Eng1bvn",
            engagementstype_titel="test1",
            primærtype_titel="?",
            bruger_uuid="b1",
            enhed_uuid="LE2",
            stillingsbetegnelse_titel="tester1",
        )
        self.session.add(engagement)
        bruger = Bruger(
            fornavn="fornavn2",
            efternavn="efternavn2",
            uuid="b2",
            bvn="b2bvn",
            cpr="cpr2",
        )
        self.session.add(bruger)
        tilknytning = Tilknytning(
            uuid="t2",
            bvn="t2bvn",
            bruger_uuid="b2",
            enhed_uuid="E3",
            tilknytningstype_titel="titel2",
            startdato="2023-10-12",
            slutdato="2030-01-02",
        )
        self.session.add(tilknytning)
        engagement = Engagement(
            uuid="Eng2",
            bvn="Eng2bvn",
            engagementstype_titel="test2",
            primærtype_titel="?",
            bruger_uuid="b2",
            enhed_uuid="LE2",
            stillingsbetegnelse_titel="tester2",
        )
        self.session.add(engagement)
        adresse = Adresse(
            uuid="A1",
            bruger_uuid="b1",
            adressetype_scope="EMAIL",
            adressetype_bvn="Email",
            adressetype_titel="Email",
            værdi="test@email.dk",
            synlighed_titel="Hemmelig",
        )
        self.session.add(adresse)
        adresse = Adresse(
            uuid="A1",
            bruger_uuid="b1",
            adressetype_scope="EMAIL",
            adressetype_bvn="AD-Email",
            adressetype_titel="AD-Email",
            værdi="AD-email@email.dk",
            synlighed_titel="Offentlig",
        )
        self.session.add(adresse)
        adresse = Adresse(
            uuid="A2",
            bruger_uuid="b1",
            adressetype_scope="PHONE",
            adressetype_bvn="AD-Telefonnummer",
            adressetype_titel="AD-Telefonnummer",
            værdi="12345678",
            synlighed_titel="",
        )
        self.session.add(adresse)
        self.session.commit()

    def tearDown(self):
        Base.metadata.drop_all(self.engine)

    # dynamic classes are fetched from graphql, here we just mock the return to check the resulting list.
    @patch(
        "reports.query_actualstate.fetch_dynamic_class",
        return_value={"t1": "Tilknytningsorganisation"},
    )
    def test_MED_data(self, _):
        # hoved_enhed = self.session.query(Enhed).all()
        # "data" comes from this class' own self.session - a sessionmaker made with SQLAlchemy.
        data = list_MED_members(self.session, {"løn": "LØN-org", "MED": "Hoved-MED"})
        self.assertEqual(
            tuple(data[0]),
            (
                "Tilknytningens startdato",
                "Tilknytningens slutdato",
                "Navn",
                "Email",
                "Hovedorganisation / Faglig organisation",
                "Telefonnummer",
                "Tilknytningstype",
                "Tilknytningsenhed",
                "Ansættelsesenhed",
                "Enhed 1",
                "Enhed 2",
            ),
        )

        self.assertEqual(
            tuple(data[1]),
            (
                "2023-01-01",
                "2023-01-02",
                "fornavn efternavn",
                "AD-email@email.dk",
                "Tilknytningsorganisation",
                "12345678",
                "titel",
                "Under-MED",
                "Under-Enhed",
                "LØN-org",
                "Under-Enhed",
            ),
        )

        self.assertEqual(
            tuple(data[2]),
            (
                "2023-10-12",
                "2030-01-02",
                "fornavn2 efternavn2",
                None,
                None,
                None,
                "titel2",
                "Under-under-MED",
                "Under-Enhed",
                "LØN-org",
                "Under-Enhed",
            ),
        )

    def test_set_of_org_units(self):
        alle_enheder = set_of_org_units(self.session, "Hoved-MED")
        self.assertEqual(alle_enheder, set(["E2", "E3"]))

    def test_EMP_data(self):
        # hoved_enhed = self.session.query(Enhed).all()
        data = list_employees(self.session, "LØN-org")
        self.assertEqual(
            tuple(data[0]),
            (
                "UUID",
                "Navn",
                "CPR",
                "AD-Email",
                "AD-Telefonnummer",
                "Enhed",
                "Stilling",
                "MA-nummer",
                "Enhed 1",
                "Enhed 2",
            ),
        )
        self.assertEqual(
            tuple(data[1]),
            (
                "b1",
                "fornavn efternavn",
                "cpr1",
                "AD-email@email.dk",
                "12345678",
                "Under-Enhed",
                "tester1",
                "Eng1bvn",
                "LØN-org",
                "Under-Enhed",
            ),
        )

        self.assertEqual(
            tuple(data[2]),
            (
                "b2",
                "fornavn2 efternavn2",
                "cpr2",
                None,
                None,
                "Under-Enhed",
                "tester2",
                "Eng2bvn",
                "LØN-org",
                "Under-Enhed",
            ),
        )


if __name__ == "__main__":
    unittest.main()
