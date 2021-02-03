import tempfile
import unittest

import xlsxwriter
from openpyxl import load_workbook

from reports.XLSXExporter import XLSXExporter


class Tests_xlxs(unittest.TestCase):
    def setUp(self):
        f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self.xlsfilename = f.name
        self.data = [
            ("Navn", "Email", "Tilknytningstype", "Enhed"),
            ("Fornavn Efternavn", "email@example.com", "Formand", "Testenhed"),
        ]
        self.employees = [
            ("Navn", "Email", "Afdeling", "Stilling", "Tlf", "CPR"),
            (
                "Fornavn Efternavn",
                "email@example.com",
                "Afdeling TEST",
                "Tester",
                "00000000",
                "123456-1234",
            ),
        ]

        workbook = xlsxwriter.Workbook(self.xlsfilename)
        excel = XLSXExporter(self.xlsfilename)
        excel.add_sheet(workbook, "MED", self.data)
        excel.add_sheet(workbook, "EMP", self.employees)
        workbook.close()

    def test_read_MED(self):
        wb = load_workbook(filename=self.xlsfilename)
        ws = wb["MED"]
        # ws contains data from excel wich can be accessed with ws.colums
        # first cell in each column contains a header, second contains data
        header = [i[0].value for i in ws.columns]
        content = [i[1].value for i in ws.columns]
        self.assertEqual(header, ["Navn", "Email", "Tilknytningstype", "Enhed"])
        self.assertEqual(
            content,
            ["Fornavn Efternavn", "email@example.com", "Formand", "Testenhed"],
        )

    def test_read_EMP(self):
        wb = load_workbook(filename=self.xlsfilename)
        ws = wb["EMP"]
        # ws contains data from excel wich can be accessed with ws.colums
        # first cell in each column contains a header, second contains data
        header = [i[0].value for i in ws.columns]
        content = [i[1].value for i in ws.columns]
        self.assertEqual(
            header, ["Navn", "Email", "Afdeling", "Stilling", "Tlf", "CPR"]
        )
        self.assertEqual(
            content,
            [
                "Fornavn Efternavn",
                "email@example.com",
                "Afdeling TEST",
                "Tester",
                "00000000",
                "123456-1234",
            ],
        )
